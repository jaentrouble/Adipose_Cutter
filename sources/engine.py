import numpy as np
from multiprocessing import Process, Queue
from PIL import Image
from .common.constants import *
from skimage import draw
from openpyxl import load_workbook
import os

# To limit loop rate
from pygame.time import Clock

class Engine(Process):
    """
    Main process that calculates all the necessary computations
    """
    # If the image is not updated, check if self._updated is switched to True
    def __init__(self, to_EngineQ:Queue, to_ConsoleQ:Queue,
                 imageQ:Queue, eventQ:Queue, etcQ:Queue):
        super().__init__(daemon=True)
        # Initial image and mask
        self.image = np.zeros((300,300,3), dtype=np.uint8)
        self.set_empty_mask()
        # Queues
        self._to_EngineQ = to_EngineQ
        self._to_ConsoleQ = to_ConsoleQ
        self._imageQ = imageQ
        self._eventQ = eventQ
        self._etcQ = etcQ
        # Modes about sending images to Viewer
        self._mask_mode = False
        self._updated = True
        # Do only one thing at a time - Do not make multiple modes
        self._mode = None
        # Initial membrane and cell color (NOT MASK COLORS)
        self.mem_color = MEMBRANE
        self.cell_color = CELL
        # Box layers
        self._box_layers = []
        self._box_start_pos = None
        # Clipped mode
        self._clipped_mode = False
        # (Color_of_layer(R,G,B), Bool mask(Width, Height, 1))
        self._layers = []
        self._cell_layers = []
        self._cell_counts = []
        self._always_on_layers = []
        self._is_drawing = False
        self._line_start_pos = None
        self._show_box = True
        # Modes related to filling
        # Ratio = (micrometer / pixel)**2  -> Because it's area ratio
        self._mp_ratio = DEFAULT_MP_RATIO
        self._mp_ratio_pixel = DEFAULT_MP_PIXEL
        self._mp_ratio_micrometer = DEFAULT_MP_MICRO

    @property
    def image(self):
        """
        This is a base image. Do not directly modify this
        """
        return self._image.copy()

    @image.setter
    def image(self, image:np.array):
        """
        Must be a shape of (width, height, 3)
        """
        if len(image.shape) != 3 and image.shape[2] != 3:
            raise TypeError('Inappropriate shape of image')
        self._image = image.astype(np.uint8)
        self._shape = self._image.shape

    @property
    def shape(self):
        """
        Do not implement shape.setter
        Shape is dependent to image and should only be setted with image
        """
        return self._shape

    @property
    def mask(self):
        """
        This is the mask on which engine computes
        """
        return self._mask.copy()

    @mask.setter
    def mask(self, mask:np.array):
        """
        Must be a shape of (width, height, 3) and same as current image
        """
        if mask.shape != self.shape:
            raise TypeError('Inappropriate shape of mask')
        self._mask = mask.astype(np.uint8)

    @property
    def cell_color(self):
        return self._cell_color

    @cell_color.setter
    def cell_color(self, color):
        if len(color) != 3 :
            raise TypeError('Wrong color given')
        self._cell_color = color

    @property
    def mem_color(self):
        return self._mem_color

    @mem_color.setter
    def mem_color(self, color):
        if len(color) != 3 :
            raise TypeError('Wrong color given')
        self._mem_color = color

    @property
    def mask_mode(self):
        return self._mask_mode

    @mask_mode.setter
    def mask_mode(self, mask_mode:bool):
        self._mask_mode = mask_mode
        self._updated = True

    @property
    def mode(self):
        return self._mode

    @mode.setter
    def mode(self, mode):
        if self._mode == MODE_DRAW_MEM:
            self.draw_stop()
        elif self._mode == MODE_FILL_MP_RATIO and self._is_drawing:
            self.fill_ratio_cancel()
        elif self._mode == MODE_DRAW_BOX:
            self.draw_box_stop()
        self._mode = mode
        self._updated = True

    def reset(self):
        self._layers = []
        self._cell_layers = []
        self._cell_counts = []
        self._always_on_layers = []
        self._is_drawing = False
        self._line_start_pos = None
        self._box_start_pos = None
        self.mode = None
        self._mask_mode = False
        self._updated = True

    def load_image(self, path:str):
        #TODO: Resize image?
        im = Image.open(path).resize((800,600))
        self.image = np.asarray(im).swapaxes(0,1)
        self.set_empty_mask()
        self.reset()
        self._updated = True

    def set_empty_mask(self):
        """
        Set a new empty mask that is the same shape as current image
        """
        self.mask = np.zeros_like(self.image)
        self._updated = True

# TODO : Change this to tf- dependent
    def set_new_mask(self, ratio:float):
        """
        ratio : if ratio * dist_to_memcolor > (1-ratio) * dist_to_cellcolor,
        the pixel is considered as cell
        
        ** This will reset all layers
        """
        ratio /=100
        dist_to_memcolor=((self.image.astype(np.int) - self.mem_color)**2).sum(
                                                        axis=2,
                                                        keepdims=True)
        dist_to_cellcolor=((self.image.astype(np.int) - self.cell_color)**2).sum(
                                                        axis=2,
                                                        keepdims=True)
        mask_bool = (dist_to_memcolor*ratio) > (dist_to_cellcolor*(1-ratio))
        self.mask = mask_bool * CELL
        self._tmp_mask = self.mask
        self.mode = None
        self._layers = []
        self._updated = True
    
    def put_image(self):
        if self._clipped_mode :
            if self._mask_mode:
                # To make drawing cell mode faster, fix other layers temporally
                if self.mode == MODE_DRAW_CELL and self._is_drawing:
                    tmp_draw_mask = self._tmp_mask.copy()
                    for c, m in self._layers:
                        np.multiply(tmp_draw_mask, np.logical_not(m), out=tmp_draw_mask)
                        np.add(tmp_draw_mask, m * np.array(c,np.uint8), out=tmp_draw_mask)
                    self._imageQ.put(tmp_draw_mask)
                else :
                    self._tmp_mask = self.mask
                    for c, m in self._layers:
                        np.multiply(self._tmp_mask, np.logical_not(m), out=self._tmp_mask)
                        np.add(self._tmp_mask, m * np.array(c,np.uint8), out=self._tmp_mask)
                    for c, m in self._cell_layers:
                        np.multiply(self._tmp_mask, np.logical_not(m), out=self._tmp_mask)
                        np.add(self._tmp_mask, m * np.array(c,np.uint8), out=self._tmp_mask)
                    for c, m in self._always_on_layers:
                        np.multiply(self._tmp_mask, np.logical_not(m), out=self._tmp_mask)
                        np.add(self._tmp_mask, m * np.array(c,np.uint8), out=self._tmp_mask)
                    self._imageQ.put(self._tmp_mask)
            else:
                tmp_image = self.image
                for c, m in self._always_on_layers:
                    np.multiply(tmp_image, np.logical_not(m), out=tmp_image)
                    np.add(tmp_image, m * np.array(c,np.uint8), out=tmp_image)
                self._imageQ.put(tmp_image)
        else :
            tmp_image = self.image
            if self._show_box:
                for c, m in self._always_on_layers:
                    np.multiply(tmp_image, np.logical_not(m), out=tmp_image)
                    np.add(tmp_image, m * np.array(c,np.uint8), out=tmp_image)
                for c, m in self._box_layers:
                    np.multiply(tmp_image, np.logical_not(m), out=tmp_image)
                    np.add(tmp_image, m * np.array(c,np.uint8), out=tmp_image)
            for c, m in self._always_on_layers:
                np.multiply(tmp_image, np.logical_not(m), out=tmp_image)
                np.add(tmp_image, m * np.array(c,np.uint8), out=tmp_image)
            self._imageQ.put(tmp_image)

    def put_mode(self):
        if self.mode != None:
            self._to_ConsoleQ.put({self.mode:None})
        else:
            self._to_ConsoleQ.put({MODE_NONE:None})
        self._to_ConsoleQ.put({FILL_MP_RATIO:self._mp_ratio})

    def put_ratio_list(self):
        self._mp_ratio = (self._mp_ratio_micrometer/self._mp_ratio_pixel)
        area_list = np.multiply(self._cell_counts, self._mp_ratio).tolist()
        self._to_ConsoleQ.put({FILL_LIST:area_list})

    # def set_mem_color(self, pos):
    #     x,y = pos
    #     new_color = self.image[x-2:x+3,y-2:y+3].mean(axis=(0,1)).astype(np.uint8)
    #     self.mem_color = new_color

    # def set_cell_color(self, pos):
    #     x,y = pos
    #     new_color = self.image[x-2:x+3,y-2:y+3].mean(axis=(0,1)).astype(np.uint8)
    #     self.cell_color = new_color

    def draw_box_start(self, pos):
        """
        Make a new layer and draw initial point (Red dot)
        """
        new_layer = np.zeros((self.shape[0],self.shape[1],1),
                             dtype=np.bool)
        color = BOX_START
        x, y = pos
        new_layer[x:x+3, y:y+3] = True
        self._always_on_layers.append((color, new_layer))
        self._box_start_pos = pos
        self._is_drawing = True
        self._updated = True

    def draw_box_stop(self):
        """
        When drawing box is interrupted.
        """
        self._box_start_pos = None
        if self._is_drawing:
            self._always_on_layers.pop()
        self._is_drawing = False
        self._etcQ.put({CROSS_CURSOR_OFF:None})
        self._updated = True


    def draw_box_end(self, pos):
        """
        Draw Box
        """
        _, last_layer = self._always_on_layers.pop()
        color = BOX_COLOR
        new_layer = np.zeros_like(last_layer)
        del last_layer
        x0, y0 = self._box_start_pos
        x1, y1 = pos
        r0, c0 = min(x0, x1), min(y0, y1)
        r1, c1 = max(x0, x1), max(y0, y1)
        new_layer[r0:r1+1,c0] = True
        new_layer[r0:r1+1,c1] = True
        new_layer[r0,c0:c1+1] = True
        new_layer[r1,c0:c1+1] = True
        self._box_layers.append((color, new_layer))
        self._box_start_pos = None
        self._is_drawing = False
        self.mode = None
        self._etcQ.put({CROSS_CURSOR_OFF:None})
        self._show_box = True
        self._updated = True

    def draw_mem_start(self, pos):
        """
        Make a new layer and draw inital point (Red dot)
        """
        new_layer = np.zeros((self.shape[0],self.shape[1],1),
                             dtype=np.bool)
        color = LINE_START
        x, y = pos
        new_layer[x:x+3, y:y+3] = True
        self._layers.append((color, new_layer))
        self._line_start_pos = pos
        self._is_drawing = True
        self._updated = True

    def draw_mem_end(self, pos):
        """
        Draw the line and start next line
        """
        _, last_layer = self._layers.pop()
        new_layer = np.zeros_like(last_layer)
        color = MEMBRANE
        del last_layer
        r0, c0 = self._line_start_pos
        r1, c1 = pos
        rr, cc, _ = draw.line_aa(r0, c0, r1, c1)
        new_layer[rr, cc] = True
        self._layers.append((color, new_layer))
        self._line_start_pos = None
        self.draw_mem_start(pos)
        self._updated = True
        
    def draw_stop(self):
        """
        Stop connecting lines; Similar to draw_undo
        """
        if self._is_drawing :
            self._layers.pop()
            self._is_drawing = False
            self._line_start_pos = None
            self._updated=True

    def draw_apply(self):
        if self._is_drawing:
            self.draw_stop()
        tmp_mask = self.mask
        for c, m in self._layers:
            np.multiply(tmp_mask, np.logical_not(m), out=tmp_mask)
            np.add(tmp_mask, m * np.array(c,np.uint8), out=tmp_mask)
        self._layers = []
        self.mask = tmp_mask
        self._updated = True

    def draw_undo(self):
        if len(self._layers)>0 :
            self._layers.pop()
            self._is_drawing = False
            self._line_start_pos = None
            self._updated = True

    def draw_cancel(self):
        self._layers = []
        self._is_drawing = False
        self._line_start_pos = None
        self.mode = None
        self._updated = True

    def draw_cell_mode_init(self):
        """
        When draw_cell mode started, not when clicked
        Call once
        """
        self._is_drawing = False
        self._updated = True


    def draw_cell_start(self, pos):
        new_layer = np.zeros((self.shape[0],self.shape[1],1),
                             dtype=np.bool)
        color = CELL
        new_layer[pos[0]-5:pos[0]+6,pos[1]-5:pos[1]+6] = True
        self._layers.append((color, new_layer))
        self._is_drawing = True
        self._updated = True
        self._etcQ.put({BIG_CURSOR_ON:None})
        self._etcQ.put({MOUSEPOS_ON:None})

    def draw_cell_continue(self, pos):
        _, last_layer = self._layers[-1]
        last_layer[pos[0]-5:pos[0]+6,pos[1]-5:pos[1]+6] = True
        self._updated = True

    def draw_cell_end(self):
        self._is_drawing = False
        self._updated = True
        self._etcQ.put({BIG_CURSOR_OFF:None})
        self._etcQ.put({MOUSEPOS_OFF:None})

    def fill_cell(self, pos):
        new_layer = np.zeros((self.shape[0],self.shape[1],1),
                             dtype=np.bool)
        pos_stack = [pos]
        mask = self.mask
        pix_count = 0
        while len(pos_stack) > 0:
            x, y = pos_stack.pop()
            while (mask[x,y] == CELL).all() and x>=0:
                x -= 1
            x += 1
            above, below = False, False
            while x < self.shape[0] and (mask[x,y]==CELL).all():
                mask[x,y] = COUNT
                new_layer[x,y] = True
                pix_count += 1
                if (not above) and (y>0) and (mask[x,y-1]==CELL).all():
                    pos_stack.append([x,y-1])
                    above = True
                elif (above) and (y>0) and (mask[x,y-1]!=CELL).all():
                    above = False
                elif (not below) and (y<self.shape[1]-1) and (mask[x,y+1]==CELL).all():
                    pos_stack.append([x,y+1])
                    below = True
                elif (below) and (y<self.shape[1]-1) and (mask[x,y+1]==CELL).all():
                    below = False
                x += 1
        self._cell_layers.append((COUNT, new_layer))
        self._cell_counts.append(pix_count)
        self._updated = True
    
    def fill_undo(self):
        if len(self._cell_layers) > 0: 
            self._cell_layers.pop()
            self._cell_counts.pop()
            self._updated = True

    def fill_ratio_start(self, pos):
        new_layer = np.zeros((self.shape[0],self.shape[1],1),
                             dtype=np.bool)
        color = LINE_START
        x, y = pos
        new_layer[x-2:x+3, y-2:y+3] = True
        self._always_on_layers.append((color, new_layer))
        self._mp_ratio_start_pos = pos
        self._is_drawing = True
        self._updated = True

    def fill_ratio_end(self, pos):
        pixel_dist = np.sqrt(np.sum(np.subtract(self._mp_ratio_start_pos,pos)**2))
        self._mp_ratio_pixel = pixel_dist
        self._always_on_layers.pop()
        self._is_drawing = False
        self._mp_ratio_start_pos = None
        self.mode = None
        self._updated = True
    
    def fill_ratio_cancel(self):
        self._mp_ratio_start_pos = None
        self._is_drawing = False
        self._always_on_layers = []
        self._updated = True

    def fill_delete(self, indices):
        for idx in indices:
            self._cell_layers.pop(idx)
            self._cell_counts.pop(idx)
        self._updated = True

    def fill_save(self, excel_dir, image_name, image_folder):
        try :
            wb = load_workbook(excel_dir)
        except :
            self._to_ConsoleQ.put({MESSAGE_BOX:'Cannot open Workbook!'})
            return
        ws = wb.worksheets[0]
        row, col = 1, 1
        while ws.cell(row, col).value != None:
            col += 1
        for area in self._cell_counts:
            ws.cell(row, col).value = area * self._mp_ratio
            ws.cell(row, col+1).value = image_name
            row += 1
        try:
            wb.save(excel_dir)
        except:
            self._to_ConsoleQ.put({MESSAGE_BOX:'Failed to Save'})
            return
        else:
            self._to_ConsoleQ.put({MESSAGE_BOX:'Saved Successfully.'\
                '\nDon\'t forget to check.'})
        # Saving the mask image
        mask_save = Image.fromarray(self._tmp_mask.swapaxes(0,1))
        new_name = image_name + '_mask.png'
        save_folder = os.path.join(image_folder,'save')
        if not os.path.exists(save_folder):
            os.mkdir(save_folder)
        filename = os.path.join(save_folder, new_name)
        mask_save.save(filename)


    def run(self):
        mainloop = True
        self._clock = Clock()
        while mainloop:
            self._clock.tick(60)
            if not self._to_EngineQ.empty():
                q = self._to_EngineQ.get()
                for k, v in q.items():
                    if k == TERMINATE:
                        mainloop = False
                    # Loading & Showing modes
                    elif k == NEWIMAGE:
                        self.load_image(v)
                    elif k == NEWMASK:
                        self.set_new_mask(*v)
                    elif k == MODE_IMAGE:
                        self.mask_mode = False
                    elif k == MODE_MASK:
                        self.mask_mode = True
                    # # Set colors & ratio
                    # elif k == SET_MEM:
                    #     self.mode = MODE_SET_MEM
                    # elif k == SET_CELL:
                    #     self.mode = MODE_SET_CELL
                    # Box Drawing
                    elif k == DRAW_BOX:
                        self.mode = MODE_DRAW_BOX
                        self._etcQ.put({CROSS_CURSOR_ON:None})
                        self._updated = True
                    elif k == SET_RATIO:
                        self.mask_mode = True
                        self.set_new_mask(v)
                    elif k == MODE_SHOW_BOX:
                        self._show_box = True
                        self._updated = True
                    elif k == MODE_HIDE_BOX:
                        self._show_box = False
                        self._updated = True
                    #Drawing modes
                    elif k == DRAW_MEM:
                        self.mode = MODE_DRAW_MEM
                        self._updated = True
                    elif k == DRAW_CELL:
                        self.mode = MODE_DRAW_CELL
                        self.draw_cell_mode_init()
                    elif k == DRAW_OFF:
                        self.draw_apply()
                    elif k == DRAW_CANCEL:
                        self.draw_cancel()
                    #Counting modes
                    elif k == FILL_CELL:
                        self.mode = MODE_FILL_CELL
                        self.draw_apply()
                    elif k == FILL_MP_RATIO:
                        self.mode = MODE_FILL_MP_RATIO
                        self._updated = True
                    elif k == FILL_DELETE:
                        self.fill_delete(v)
                    elif k == FILL_SAVE:
                        self.fill_save(*v)
                    elif k == FILL_MICRO:
                        self._mp_ratio_micrometer = v
                        self._updated = True

            if not self._eventQ.empty():
                q = self._eventQ.get()
                for k,v in q.items():
                    if k == MOUSEDOWN:
                        # v : mouse pos which came from Viewer
                        # # Set color
                        # if self.mode == MODE_SET_MEM:
                        #     self.set_mem_color(v)
                        #     self._color_mode = None
                        #     self._to_ConsoleQ.put({SET_MEM:self.mem_color})
                        # elif self.mode == MODE_SET_CELL:
                        #     self.set_cell_color(v)
                        #     self._color_mode = None
                        #     self._to_ConsoleQ.put({SET_CELL:self.cell_color})
                        # Box mode
                        if self.mode == MODE_DRAW_BOX:
                            if not self._is_drawing:
                                self.draw_box_start(v)
                            else:
                                self.draw_box_end(v)
                        # Drawing mode
                        if self.mode == MODE_DRAW_MEM:
                            if not self._is_drawing:
                                self.draw_mem_start(v)
                            else:
                                self.draw_mem_end(v)
                        elif self.mode == MODE_DRAW_CELL:
                            self.draw_cell_start(v)
                        # Counting mode
                        elif self.mode == MODE_FILL_CELL:
                            self.fill_cell(v)
                        elif self.mode == MODE_FILL_MP_RATIO:
                            if not self._is_drawing:
                                self.fill_ratio_start(v)
                            else :
                                self.fill_ratio_end(v)
                    elif k == MOUSEDOWN_RIGHT:
                        self.draw_stop()
                    elif k == MOUSEUP:
                        if self.mode == MODE_DRAW_CELL:
                            self.draw_cell_end()
                    elif k == MOUSEPOS:
                        if self.mode == MODE_DRAW_CELL:
                            if self._is_drawing:
                                self.draw_cell_continue(v)
                    # Keyboard events
                    elif k == K_Z:
                        if self.mode == MODE_DRAW_CELL or\
                            self.mode == MODE_DRAW_MEM:
                            self.draw_undo()
                        elif self.mode == MODE_FILL_CELL:
                            self.fill_undo()
                    elif k == K_ENTER:
                        if self.mode == MODE_DRAW_MEM or\
                        self.mode == MODE_DRAW_CELL:
                            self.draw_apply()

            if self._updated:
                self.put_image()
                self.put_ratio_list()
                self.put_mode()
                self._updated = False